import pandas as pd
import openpyxl
import os
import re
import unicodedata

def normalizar_extremo(texto):
    """Limpia nombres que vienen con formato de archivo (1. NOMBRE-ICAT.pdf)"""
    if pd.isna(texto): return ""
    s = str(texto).upper().strip()
    # Limpieza específica de basura de archivos
    s = s.replace('.PDF', '').replace('-ICAT PUEBLA', '').replace('ICAT PUEBLA', '')
    s = re.sub(r'^\d+[\s.]*', '', s)
    # Quitar acentos
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    # Dejar solo letras y Ñ
    s = re.sub(r'[^A-ZÑ]', ' ', s)
    return set(s.split())

def leer_maestra_inteligente(ruta):
    """Busca los datos saltando posibles filas de logotipos."""
    for i in range(15):
        try:
            df = pd.read_excel(ruta, skiprows=i)
            if any("NOMBRE" in str(c).upper() for c in df.columns):
                return df
        except: continue
    return pd.read_excel(ruta)

def complementar_excels(ruta_formato_parcial, ruta_maestro, nombre_salida):
    try:
        # 1. Leer Base de Datos Maestra
        df_mae = leer_maestra_inteligente(ruta_maestro)

        # 2. Identificar columnas clave en la Maestra de forma dinámica
        def buscar_col(keywords):
            for c in df_mae.columns:
                if any(k in str(c).upper() for k in keywords): return c
            return None

        c_nom = buscar_col(["NOMBRE", "ALUMNO", "INSCRITA"])
        c_sex = buscar_col(["SEXO", "GENERO"])
        c_eda = buscar_col(["EDAD"])
        c_mun = buscar_col(["MUNICIPIO", "PROCEDENCIA"])
        c_est = buscar_col(["ESTATUS", "SITUACION"])
        c_ins = buscar_col(["INSTRUCTOR", "MAESTRO", "PRESTADOR"])
        c_cur = buscar_col(["CURSO", "ESTANDAR", "COMPETENCIA"])
        # NUEVO: Columna UNIDAD en la DB
        c_uni = buscar_col(["UNIDAD", "CENTRO", "AULA"])

        # 3. Cargar datos en memoria
        maestra_data = []
        for _, fila in df_mae.iterrows():
            if pd.notna(fila[c_nom]):
                maestra_data.append({
                    'palabras': normalizar_extremo(fila[c_nom]),
                    'info': {
                        'SEXO': fila[c_sex] if c_sex else "",
                        'EDAD': fila[c_eda] if c_eda else "",
                        'MUNICIPIO': fila[c_mun] if c_mun else "",
                        'ESTATUS': fila[c_est] if c_est else "",
                        'INSTRUCTOR': fila[c_ins] if c_ins else "",
                        'CURSO': fila[c_cur] if c_cur else "",
                        'UNIDAD': fila[c_uni] if c_uni else "" # Guardamos la Unidad
                    }
                })

        # 4. Abrir Formato a completar
        wb = openpyxl.load_workbook(ruta_formato_parcial)
        sheet = wb.active
        encontrados = 0

        # 5. Llenado por celdas (Paso 2)
        # Basado en tu imagen: B=Nombre, C=Sexo, D=Edad, E=Municipio, F=Estatus, G=Curso, I=Prestador, J=Unidad
        for r in range(6, sheet.max_row + 1):
            valor_celda = sheet.cell(row=r, column=2).value
            if not valor_celda: continue
            
            palabras_f = normalizar_extremo(valor_celda)
            if not palabras_f: continue

            for persona in maestra_data:
                # Si el nombre coincide (ignorando orden de apellidos y basura de archivos)
                if palabras_f.issubset(persona['palabras']) or persona['palabras'].issubset(palabras_f):
                    d = persona['info']
                    
                    sheet.cell(row=r, column=3).value = d['SEXO']       # C
                    sheet.cell(row=r, column=4).value = d['EDAD']       # D
                    sheet.cell(row=r, column=5).value = d['MUNICIPIO']  # E
                    sheet.cell(row=r, column=6).value = d['ESTATUS']    # F
                    sheet.cell(row=r, column=7).value = d['CURSO']      # G
                    sheet.cell(row=r, column=9).value = d['INSTRUCTOR'] # I
                    sheet.cell(row=r, column=10).value = d['UNIDAD']    # J (NUEVA COLUMNA)
                    
                    encontrados += 1
                    break
        
        print(f"-> ÉXITO: Se vincularon {encontrados} alumnos con Unidad incluida.")

        # 6. Guardado del reporte final
        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs')
        if not os.path.exists(out_dir): os.makedirs(out_dir)
        ruta_final = os.path.join(out_dir, nombre_salida)
        wb.save(ruta_final)
        wb.close()
        return ruta_final

    except Exception as e:
        print(f"Error técnico en el Merger: {e}")
        return None