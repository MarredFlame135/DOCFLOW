import openpyxl
import os
import re
from openpyxl.cell.cell import MergedCell

def limpiar_texto_excel(texto):
    if not texto: return ""
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\xff]', '', str(texto)).strip()

def descombinar_celda(sheet, row, col):
    """Si una celda está combinada, la descombina para poder escribir en ella."""
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in list(sheet.merged_cells.ranges):
            if cell.coordinate in merged_range:
                sheet.unmerge_cells(str(merged_range))
                break

def llenar_excel(lista_datos, ruta_plantilla, nombre_salida):
    if not os.path.exists(ruta_plantilla):
        raise FileNotFoundError(f"No encontré la plantilla en: {ruta_plantilla}")

    wb = openpyxl.load_workbook(ruta_plantilla)
    sheet = wb.active 
    fila = 6 # Fila de inicio

    for dato in lista_datos:
        # 1. Obtenemos los datos
        nombre = limpiar_texto_excel(dato.get("Nombre", "N/A")).upper()
        curso = limpiar_texto_excel(dato.get("Curso", "N/A"))
        fecha = limpiar_texto_excel(dato.get("Fecha", "N/A"))

        # 2. Verificamos y descombinamos antes de escribir (Columnas 2, 7 y 8)
        for col_index in [2, 7, 8]:
            descombinar_celda(sheet, fila, col_index)

        # 3. Ahora escribimos con seguridad
        sheet.cell(row=fila, column=2).value = nombre
        sheet.cell(row=fila, column=7).value = curso
        sheet.cell(row=fila, column=8).value = fecha
        
        fila += 1

    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs')
    if not os.path.exists(out_dir): os.makedirs(out_dir)
    ruta_final = os.path.join(out_dir, nombre_salida)

    wb.save(ruta_final)
    wb.close()
    return ruta_final